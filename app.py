import logging
import asyncio
import time
import html
import csv
import os
import sqlite3
import requests
import aiohttp
import random
import string
from datetime import datetime, timedelta
from io import StringIO, BytesIO
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.firefox.service import Service
from selenium.webdriver.firefox.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.firefox import GeckoDriverManager
from pytube import YouTube, Search
from aiogram import Bot, Dispatcher, executor, types
from aiogram.types import InlineKeyboardMarkup, InlineKeyboardButton, ReplyKeyboardRemove
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
import json

# Импорт парсеров из нового модуля
from parsers import ParserWB, HHParserApp, YouTubeParser, WeatherParser

# Настройка логирования
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)
#━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# Токены
# Загрузка токенов из config.json
with open('config.json', 'r', encoding='utf-8') as f:
    config = json.load(f)
API_TOKEN = config.get('API_TOKEN')
CRYPTO_PAY_API_TOKEN = config.get('CRYPTO_PAY_API_TOKEN')
#━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# Инициализация бота и диспетчера
bot = Bot(token=API_TOKEN)
dp = Dispatcher(bot)
#━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# Глобальные словари для хранения состояний пользователей и результатов запросов
user_states = {}       # {user_id: {"action": <тип запроса>}}
user_last_data = {}    # {user_id: { 'action': str, 'data': структурированные данные, 'text': результат в виде строки, 'hh_params': {...} }}
#━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# Глобальные переменные для пагинации
pagination_sessions = {}  # ключ: session_id, значение: dict с страницами, текущей страницей, chat_id и message_id
pagination_counter = 0    # для генерации уникальных session_id
#━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# Функция для разбивки сообщений на части (макс. 4096 символов)
def split_message(message: str):
    max_message_length = 4096
    return [message[i:i + max_message_length] for i in range(0, len(message), max_message_length)]
#━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# Функция для формирования inline‑клавиатуры пагинации
def build_pagination_keyboard(session_id, current_page, total_pages, extra_buttons=None):
    keyboard = InlineKeyboardMarkup(row_width=3)
    buttons = []
    
    # Добавляем кнопки навигации только если страниц больше одной
    if total_pages > 1:
        if current_page > 0:
            buttons.append(InlineKeyboardButton("⬅️", callback_data=f"paginate:{session_id}:{current_page - 1}"))
        
        # Добавляем номер текущей страницы
        buttons.append(InlineKeyboardButton(f"{current_page + 1}/{total_pages}", callback_data="ignore"))
        
        if current_page < total_pages - 1:
            buttons.append(InlineKeyboardButton("➡️", callback_data=f"paginate:{session_id}:{current_page + 1}"))
        
        # Добавляем кнопки в клавиатуру
        if buttons:
            keyboard.row(*buttons)
    
    # Добавляем дополнительные кнопки (например, для скачивания) в любом случае
    if extra_buttons:
        for button in extra_buttons:
            keyboard.row(button)
    
    return keyboard
#━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# Функция для отправки ответа с пагинацией, если страниц более одной
async def send_paginated_message(chat_id, pages, extra_buttons=None):
    global pagination_counter
    session_id = f"p_{pagination_counter}"
    pagination_counter += 1
    keyboard = build_pagination_keyboard(session_id, 0, len(pages), extra_buttons)
    msg = await bot.send_message(chat_id, pages[0], parse_mode='HTML', reply_markup=keyboard)
    pagination_sessions[session_id] = {
         'pages': pages,
         'current_page': 0,
         'chat_id': chat_id,
         'message_id': msg.message_id,
         'extra_buttons': extra_buttons
    }
#━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# Универсальная функция для отправки ответа: если несколько страниц – с пагинацией, иначе обычное сообщение
async def send_response(chat_id, text, add_download_buttons=False, action=None, extra_buttons=None):
    pages = split_message(text)
    if len(pages) > 1:
        await send_paginated_message(chat_id, pages, extra_buttons)
    else:
        await bot.send_message(chat_id, pages[0], parse_mode='HTML', reply_markup=build_pagination_keyboard("dummy", 0, 1, extra_buttons))
    # Если нужно добавить кнопки для скачивания файла с результатами
    if add_download_buttons and action:
        keyboard = get_download_keyboard(action)
        await bot.send_message(chat_id, "Выберите формат для выгрузки результата:", reply_markup=keyboard)
#━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# Обработчик callback для пагинации
@dp.callback_query_handler(lambda c: c.data.startswith("paginate:") or c.data == "ignore")
async def process_pagination_callback(callback_query: types.CallbackQuery):
    try:
        # Если нажата кнопка с номером страницы, игнорируем
        if callback_query.data == "ignore":
            await callback_query.answer()
            return

        data = callback_query.data.split(":")  # формат: paginate:session_id:new_page
        if len(data) != 3:
            await callback_query.answer("Некорректные данные.")
            return

        session_id = data[1]
        try:
            new_page = int(data[2])
        except:
            await callback_query.answer("Некорректный номер страницы.")
            return

        session = pagination_sessions.get(session_id)
        if not session:
            await callback_query.message.edit_text(
                "❌ Сессия устарела. Пожалуйста, выполните поиск заново.",
                reply_markup=None
            )
            return

        pages = session['pages']
        if new_page < 0 or new_page >= len(pages):
            await callback_query.answer("Неверный номер страницы.")
            return

        session['current_page'] = new_page
        extra_buttons = session.get('extra_buttons')
        keyboard = build_pagination_keyboard(session_id, new_page, len(pages), extra_buttons)

        try:
            await bot.edit_message_text(
                pages[new_page],
                chat_id=session['chat_id'],
                message_id=session['message_id'],
                parse_mode='HTML',
                reply_markup=keyboard
            )
        except Exception as e:
            logger.error(f"Ошибка при обновлении сообщения: {e}")
            await callback_query.message.edit_text(
                "❌ Не удалось обновить сообщение. Пожалуйста, выполните поиск заново.",
                reply_markup=None
            )
            return

        await callback_query.answer()
    except Exception as e:
        logger.error(f"Ошибка в обработчике пагинации: {e}")
        try:
            await callback_query.message.edit_text(
                "❌ Произошла ошибка. Пожалуйста, выполните поиск заново.",
                reply_markup=None
            )
        except:
            pass
#━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# ┏━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━┓
# ┃  SQLite DB Helper                                            ┃
# ┗━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━┛
#━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
class DBHelper:
    def __init__(self, db_path='bot.db'):
        self.conn = sqlite3.connect(db_path, check_same_thread=False)
        self.create_tables()

    def create_tables(self):
        self.conn.execute("""
            CREATE TABLE IF NOT EXISTS users (
                user_id INTEGER PRIMARY KEY,
                subscription INTEGER DEFAULT 0,
                subscription_expiry TEXT,
                request_count INTEGER DEFAULT 0,
                last_reset TEXT,
                points INTEGER DEFAULT 0,
                referral_code TEXT,
                referrals INTEGER DEFAULT 0
            )
        """)
        self.conn.execute("""
            CREATE TABLE IF NOT EXISTS history (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER,
                action TEXT,
                query TEXT,
                timestamp TEXT
            )
        """)
        self.conn.commit()

    def get_user(self, user_id):
        cursor = self.conn.execute("SELECT subscription, subscription_expiry, request_count, last_reset, points, referral_code, referrals FROM users WHERE user_id=?", (user_id,))
        row = cursor.fetchone()
        return row

    def get_user_by_referral(self, ref_code):
        cursor = self.conn.execute("SELECT user_id, subscription, subscription_expiry, request_count, points, referral_code, referrals FROM users WHERE referral_code=?", (ref_code,))
        row = cursor.fetchone()
        return row

    def add_user(self, user_id):
        now = datetime.now().isoformat()
        referral_code = ''.join(random.choices(string.ascii_uppercase + string.digits, k=6))
        self.conn.execute("INSERT OR IGNORE INTO users (user_id, subscription, subscription_expiry, request_count, last_reset, points, referral_code, referrals) VALUES (?, 0, NULL, 0, ?, 0, ?, 0)", (user_id, now, referral_code))
        self.conn.commit()

    def add_referral(self, ref_code):
        self.conn.execute("UPDATE users SET referrals = referrals + 1 WHERE referral_code=?", (ref_code,))
        self.conn.commit()

    def update_subscription(self, user_id, subscription):
        if subscription == 1:
            expiry_date = (datetime.now() + timedelta(days=30)).isoformat()
        else:
            expiry_date = None
        self.conn.execute("UPDATE users SET subscription=?, subscription_expiry=? WHERE user_id=?", (subscription, expiry_date, user_id))
        self.conn.commit()

    def reset_request_count(self, user_id):
        now = datetime.now().isoformat()
        self.conn.execute("UPDATE users SET request_count=0, last_reset=? WHERE user_id=?", (now, user_id))
        self.conn.commit()

    def increment_request_count(self, user_id):
        self.conn.execute("UPDATE users SET request_count = request_count + 1, points = points + 1 WHERE user_id=?", (user_id,))
        self.conn.commit()

    def check_request_limit(self, user_id):
        self.add_user(user_id)
        user = self.get_user(user_id)
        subscription, subscription_expiry, request_count, last_reset, points, referral_code, referrals = user
        now = datetime.now()
        try:
            last_reset_dt = datetime.fromisoformat(last_reset)
        except Exception:
            last_reset_dt = now
        if (now - last_reset_dt).total_seconds() > 86400:
            self.reset_request_count(user_id)
            request_count = 0
        if subscription == 1:
            self.increment_request_count(user_id)
            return True, ""
        else:
            if request_count >= 5:
                return False, "Превышен лимит запросов (5 запросов в 24 часа). Оформите подписку через команду /subscribe."
            else:
                self.increment_request_count(user_id)
                return True, ""

    def add_history(self, user_id, action, query):
        now = datetime.now().isoformat()
        self.conn.execute("INSERT INTO history (user_id, action, query, timestamp) VALUES (?, ?, ?, ?)", (user_id, action, query, now))
        self.conn.commit()

    def get_history(self, user_id, limit=10):
        cursor = self.conn.execute("SELECT action, query, timestamp FROM history WHERE user_id=? ORDER BY id DESC LIMIT ?", (user_id, limit))
        return cursor.fetchall()

    def check_all_subscriptions(self):
        now = datetime.now().isoformat()
        self.conn.execute("UPDATE users SET subscription=0, subscription_expiry=NULL WHERE subscription=1 AND subscription_expiry < ?", (now,))
        self.conn.commit()

db_helper = DBHelper()
#━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# Фоновая задача для ежедневной проверки подписок
async def subscription_checker():
    while True:
        try:
            db_helper.check_all_subscriptions()
            logger.info("Проверка подписок выполнена.")
        except Exception as e:
            logger.error(f"Ошибка при проверке подписок: {e}")
        await asyncio.sleep(86400)  # 24 часа
#━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# ┏━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━┓
# ┃     ПАРСЕР WILDBERRIES (WB)   ┃
# ┗━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━┛
#━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
class ParserWB:
    def __init__(self):
        self.products_data = []

    def setup_driver(self):
        firefox_options = Options()
        firefox_options.add_argument("--headless")
        firefox_options.set_preference("dom.webdriver.enabled", False)
        firefox_options.set_preference("useAutomationExtension", False)
        service = Service(GeckoDriverManager().install())
        self.driver = webdriver.Firefox(service=service, options=firefox_options)

    def parse(self, query: str):
        try:
            self.setup_driver()
            url = f"https://www.wildberries.ru/catalog/0/search.aspx?search={query}"
            self.driver.get(url)
            wait = WebDriverWait(self.driver, 20)
            try:
                wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, "div.product-card__wrapper")))
            except Exception as e:
                raise Exception(f"Карточки товаров не загрузились: {e}")
            last_height = self.driver.execute_script("return document.body.scrollHeight")
            while True:
                self.driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
                time.sleep(2)
                new_height = self.driver.execute_script("return document.body.scrollHeight")
                if new_height == last_height:
                    break
                last_height = new_height
            product_cards = self.driver.find_elements(By.CSS_SELECTOR, "div.product-card__wrapper")
            self.products_data = [self.extract_product_data(card) for card in product_cards]
            self.products_data = [p for p in self.products_data if p is not None]
            return self.products_data
        finally:
            if hasattr(self, 'driver') and self.driver:
                self.driver.quit()

    def extract_product_data(self, card):
        try:
            name = card.find_element(By.CSS_SELECTOR, "span.product-card__name").text.strip()
            try:
                price_element = card.find_element(By.CSS_SELECTOR, "span.price__lower-price, ins.price__lower-price")
                price_text = price_element.text.strip().replace('\u2009', ' ')
            except:
                price_text = "0"
            try:
                rating = card.find_element(By.CSS_SELECTOR, "span.address-rate-mini").text.strip()
            except:
                rating = "0"
            try:
                link = card.find_element(By.CSS_SELECTOR, "a.product-card__link").get_attribute('href')
            except:
                link = "Ссылка не найдена"
            try:
                reviews = card.find_element(By.CSS_SELECTOR, "span.product-card__count").text.strip()
                if not reviews:
                    reviews = "0"
            except:
                reviews = "0"

            return {
                "Название": name,
                "Цена": price_text,
                "Рейтинг": rating,
                "Ссылка": link,
                "Отзывы": reviews
            }
        except Exception as e:
            logger.error(f"Ошибка при извлечении данных: {e}")
            return None
#━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# ┏━━━━━━━━━━━━━━━━━━━━━━━━━━━━━┓
# ┃         ПАРСЕР HH.ru        ┃
# ┗━━━━━━━━━━━━━━━━━━━━━━━━━━━━━┛
#━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
class HHParserApp:
    def __init__(self):
        self.status = "Готово"
        self.max_results = 20  # По умолчанию 20 вакансий

    def get_area_id(self, city_name):
        try:
            response = requests.get('https://api.hh.ru/areas')
            areas = response.json()
            for area in areas:
                for item in area['areas']:
                    if city_name.lower() in item['name'].lower():
                        return item['id']
                    for child in item['areas']:
                        if city_name.lower() in child['name'].lower():
                            return child['id']
            return None
        except Exception as e:
            logger.error(f"Ошибка: {str(e)}")
            return None

    def parse_vacancy(self, vacancy):
        salary = vacancy.get('salary')
        return {
            'название': vacancy.get('name'),
            'работодатель': vacancy.get('employer', {}).get('name'),
            'зарплата_от': salary.get('from') if salary else None,
            'зарплата_до': salary.get('to') if salary else None,
            'валюта': salary.get('currency') if salary else None,
            'опыт': vacancy.get('experience', {}).get('name'),
            'график_работы': vacancy.get('schedule', {}).get('name'),
            'url': vacancy.get('alternate_url'),
            'дата публикации': vacancy.get('published_at')
        }

    def get_vacancies(self, query, area_id, salary_filter=None, employment_filter=None, remote_filter=None):
        all_vacancies = []
        params = {
            'text': query,
            'area': area_id,
            'per_page': 100,
            'page': 0
        }
        if salary_filter:
            params['salary'] = salary_filter
        if employment_filter:
            params['employment'] = employment_filter
        if remote_filter:
            params['schedule'] = remote_filter
        try:
            while True:
                response = requests.get('https://api.hh.ru/vacancies', params=params)
                response.raise_for_status()
                data = response.json()
                vacancies = data.get('items', [])
                if not vacancies:
                    break
                for vacancy in vacancies:
                    parsed = self.parse_vacancy(vacancy)
                    all_vacancies.append(parsed)
                    # Если достигли лимита в 20 вакансий, прерываем загрузку
                    if len(all_vacancies) >= 20:
                        return all_vacancies
                logger.info(f"Обработано {len(all_vacancies)} вакансий...")
                if params['page'] >= data['pages'] - 1:
                    break
                params['page'] += 1
            return all_vacancies if all_vacancies else []
        except Exception as e:
            logger.error(f"Ошибка: {str(e)}")
            return []

    def run_parser(self, query, area_id, salary_filter=None, employment_filter=None, remote_filter=None):
        return self.get_vacancies(query, area_id, salary_filter, employment_filter, remote_filter)
#━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# ┏━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━┓
# ┃               О Б Р А Б О Т Ч И К И   К О М А Н Д            ┃
# ┃                           Б О Т А                            ┃
# ┗━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━┛                                                                                  
#━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
@dp.message_handler(commands=['start', 'help'])
async def send_welcome(message: types.Message):
    user_id = message.from_user.id
    db_helper.add_user(user_id)
    args = message.get_args().strip()
    if args:
        ref_code = args
        db_helper.add_referral(ref_code)
        ref_user = db_helper.get_user_by_referral(ref_code)
        if ref_user:
            current_referrals = ref_user[6]
            if current_referrals >= 20 and ref_user[0] and ref_user[1] == 0:
                db_helper.update_subscription(ref_user[0], 1)
                try:
                    await bot.send_message(ref_user[0], "Поздравляем! За привлечение 20 друзей вы получили подписку на месяц!")
                except Exception as e:
                    logger.error(f"Ошибка отправки уведомления о подписке: {e}")
    user = db_helper.get_user(user_id)
    subscription = user[0] if user else 0
    keyboard = InlineKeyboardMarkup()
    keyboard.add(
        InlineKeyboardButton("Youtube", callback_data="parser_yt"),
        InlineKeyboardButton("Погода", callback_data="weather")
    )
    keyboard.add(
        InlineKeyboardButton("HHru", callback_data="parse_hh"),
        InlineKeyboardButton("Wildberries", callback_data="parse_wb")
    )
    if subscription == 0:
        keyboard.add(InlineKeyboardButton("💎 Купить подписку", callback_data="subscribe"))
    keyboard.add(InlineKeyboardButton("👤 Профиль", callback_data="profile"))
    keyboard.add(InlineKeyboardButton("📜 История", callback_data="history"))
    keyboard.add(InlineKeyboardButton("🔗 Реферальная ссылка", callback_data="referral"))
    await message.reply(
        "Привет! Я бот для парсинга данных.\n\n"
        "Доступные команды:\n"
        "Wildberries - поиск товаров на Wildberries\n"
        "HHru - поиск вакансий на HH.ru\n"
        "Погода - получение информации о погоде\n"
        "Youtube - поиск видео на YouTube\n"
        "Купить подписку - оформить подписку (неограниченные запросы)\n"
        "В бесплатной версии доступно 5 запросов в 24 часа\n\n"
        "Также поддерживаются голосовые запросы, сортировка/фильтрация результатов, реферальные ссылки и начисление баллов за активность.",
        reply_markup=keyboard
    )
#━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
@dp.callback_query_handler(lambda c: c.data == "subscribe")
async def process_subscribe_callback(callback_query: types.CallbackQuery):
    await handle_subscribe(callback_query.message)
    await callback_query.answer()
#━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
@dp.callback_query_handler(lambda c: c.data == "profile")
async def process_profile_callback(callback_query: types.CallbackQuery):
    user_id = callback_query.from_user.id
    db_helper.add_user(user_id)
    user = db_helper.get_user(user_id)
    subscription, subscription_expiry, request_count, last_reset, points, referral_code, referrals = user
    now = datetime.now()
    try:
        last_reset_dt = datetime.fromisoformat(last_reset)
    except Exception:
        last_reset_dt = now
    time_since_reset = now - last_reset_dt
    seconds_left = max(0, 86400 - int(time_since_reset.total_seconds()))
    hours_left = seconds_left // 3600
    minutes_left = (seconds_left % 3600) // 60
    if subscription == 1:
        profile_text = (
            f"👤 <b>Профиль пользователя</b>\n\n"
            f"✅ Подписка активна\n"
            f"До: {subscription_expiry}\n"
            f"Запросов выполнено: {request_count} (лимит неограничен)\n"
            f"Баллы: {points}\n"
            f"Ваш реферальный код: {referral_code}\n"
            f"Рефералов: {referrals}"
        )
    else:
        profile_text = (
            f"👤 <b>Профиль пользователя</b>\n\n"
            f"❌ Подписка не оформлена\n"
            f"Запросов выполнено: {request_count}/5\n"
            f"До обновления лимита: {hours_left}ч {minutes_left}м\n"
            f"Баллы: {points}\n"
            f"Ваш реферальный код: {referral_code}\n"
            f"Рефералов: {referrals}"
        )
    await bot.send_message(callback_query.from_user.id, profile_text, parse_mode='HTML')
    await callback_query.answer()
#━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
@dp.callback_query_handler(lambda c: c.data == "history")
async def process_history_callback(callback_query: types.CallbackQuery):
    user_id = callback_query.from_user.id
    history = db_helper.get_history(user_id)
    if not history:
        text = "История запросов пуста."
    else:
        text = "📜 <b>История ваших запросов:</b>\n"
        for action, query, timestamp in history:
            text += f"[{timestamp.split('T')[0]}] {action}: {query}\n"
    await bot.send_message(user_id, text, parse_mode='HTML')
    await callback_query.answer()
#━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
@dp.callback_query_handler(lambda c: c.data == "referral")
async def process_referral_callback(callback_query: types.CallbackQuery):
    user_id = callback_query.from_user.id
    user = db_helper.get_user(user_id)
    referral_code = user[5] if user else "N/A"
    referral_link = f"https://t.me/blacklistcheckeeeeeeeeer_bot?start={referral_code}"
    text = f"🔗 Ваша реферальная ссылка:\n{referral_link}"
    await bot.send_message(user_id, text)
    await callback_query.answer()
#━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
@dp.callback_query_handler(lambda c: c.data in ["parser_yt", "weather", "parse_wb"])
async def process_callback_query_buttons(callback_query: types.CallbackQuery):
    action = callback_query.data
    user_id = callback_query.from_user.id
    user_states[user_id] = {"action": action}
    if action == "parser_yt":
        prompt = "Введите поисковый запрос для YouTube:"
    elif action == "weather":
        prompt = "Введите название города для получения прогноза погоды:"
    elif action == "parse_wb":
        prompt = "Введите поисковый запрос для поиска товаров на Wildberries:"
    await bot.send_message(user_id, prompt)
    await callback_query.answer()

@dp.message_handler(lambda message: message.from_user.id in user_states and 
                   user_states[message.from_user.id].get("action") == "parse_hh" and 
                   "state" not in user_states[message.from_user.id])
async def handle_old_hh_format(message: types.Message):
    # Перенаправляем на новый формат диалога
    user_id = message.from_user.id
    msg = await bot.send_message(
        user_id,
        "💼 Теперь поиск работы стал удобнее! Давайте начнем сначала.\n\n"
        "Какую работу ищете? Введите название профессии или должности:"
    )
    user_states[user_id] = {
        "action": "parse_hh",
        "state": "waiting_job",
        "data": {},
        "dialog_message_id": msg.message_id
    }

@dp.message_handler(lambda message: message.from_user.id in user_states and 
                   user_states[message.from_user.id].get("action") == "parse_hh" and 
                   user_states[message.from_user.id].get("state") in HH_STATES)
async def handle_hh_dialog(message: types.Message):
    user_id = message.from_user.id
    current_state = user_states[user_id]["state"]
    user_data = user_states[user_id].get("data", {})
    dialog_message_id = user_states[user_id].get("dialog_message_id")

    # Удаляем сообщение пользователя для чистоты чата
    try:
        await message.delete()
    except:
        pass

    if current_state == "waiting_job":
        user_data["job"] = message.text.strip()
        user_states[user_id]["state"] = "waiting_city"
        await bot.edit_message_text(
            "✅ Профессия: " + user_data["job"] + "\n\n"
            "В каком городе ищете работу?",
            chat_id=user_id,
            message_id=dialog_message_id
        )

    elif current_state == "waiting_city":
        user_data["city"] = message.text.strip()
        user_states[user_id]["state"] = "waiting_salary"
        await bot.edit_message_text(
            "✅ Профессия: " + user_data["job"] + "\n"
            "✅ Город: " + user_data["city"] + "\n\n"
            "Укажите желаемую зарплату (или напишите 'любая'):",
            chat_id=user_id,
            message_id=dialog_message_id
        )

    elif current_state == "waiting_salary":
        salary_text = message.text.strip().lower()
        if salary_text != "любая":
            try:
                salary = int(salary_text)
                user_data["salary"] = salary
            except ValueError:
                await bot.edit_message_text(
                    "✅ Профессия: " + user_data["job"] + "\n"
                    "✅ Город: " + user_data["city"] + "\n"
                    "❌ Ошибка: введите число или слово 'любая'\n\n"
                    "Укажите желаемую зарплату:",
                    chat_id=user_id,
                    message_id=dialog_message_id
                )
                return
        else:
            user_data["salary"] = None

        # Создаем клавиатуру для выбора формата работы
        keyboard = InlineKeyboardMarkup(row_width=1)
        keyboard.add(
            InlineKeyboardButton("🏢 Офис", callback_data="work_type:office"),
            InlineKeyboardButton("🏠 Удаленная работа", callback_data="work_type:remote"),
            InlineKeyboardButton("🔄 Гибридный формат", callback_data="work_type:hybrid"),
            InlineKeyboardButton("📋 Любой формат", callback_data="work_type:any")
        )
        user_states[user_id]["state"] = "waiting_work_type"
        
        salary_display = "любая" if user_data["salary"] is None else str(user_data["salary"])
        await bot.edit_message_text(
            "✅ Профессия: " + user_data["job"] + "\n"
            "✅ Город: " + user_data["city"] + "\n"
            "✅ Зарплата: " + salary_display + "\n\n"
            "Выберите предпочтительный формат работы:",
            chat_id=user_id,
            message_id=dialog_message_id,
            reply_markup=keyboard
        )

    user_states[user_id]["data"] = user_data

@dp.callback_query_handler(lambda c: c.data.startswith("work_type:"))
async def process_work_type(callback_query: types.CallbackQuery):
    try:
        user_id = callback_query.from_user.id
        if user_id not in user_states or user_states[user_id].get("action") != "parse_hh":
            await callback_query.message.edit_text(
                "❌ Сессия поиска устарела. Пожалуйста, начните поиск заново.",
                reply_markup=None
            )
            return

        work_type = callback_query.data.split(":")[1]
        user_data = user_states[user_id].get("data", {})
        user_data["work_type"] = work_type

        # Обновляем сообщение, показывая все выбранные параметры
        work_type_text = {
            "office": "Офис",
            "remote": "Удаленная работа",
            "hybrid": "Гибридный формат",
            "any": "Любой формат"
        }.get(work_type, "Не указан")

        salary_display = "любая" if user_data.get("salary") is None else str(user_data["salary"])
        await callback_query.message.edit_text(
            "✅ Профессия: " + user_data["job"] + "\n"
            "✅ Город: " + user_data["city"] + "\n"
            "✅ Зарплата: " + salary_display + "\n"
            "✅ Формат работы: " + work_type_text + "\n\n"
            "🔄 Начинаю поиск вакансий..."
        )

        # Начинаем поиск вакансий
        await start_hh_search(callback_query.message, user_id, user_data)
        await callback_query.answer()
    except Exception as e:
        logger.error(f"Ошибка в обработчике выбора типа работы: {e}")
        try:
            await callback_query.message.edit_text(
                "❌ Произошла ошибка. Пожалуйста, начните поиск заново.",
                reply_markup=None
            )
        except:
            pass

# Обработка inline‑кнопок для сортировки Wildberries
@dp.callback_query_handler(lambda c: c.data.startswith("sort_wb:"))
async def process_sort_wb_callback(callback_query: types.CallbackQuery):
    try:
        data = callback_query.data.split(":")  # формат: sort_wb:<field>:<order>
        if len(data) != 3:
            await callback_query.answer("Некорректные данные сортировки.")
            return

        sort_field = data[1]
        sort_order = data[2]
        user_id = callback_query.from_user.id

        if user_id not in user_last_data or user_last_data[user_id].get('action') != 'parse_wb':
            await callback_query.answer("Нет данных для сортировки.")
            return

        products = user_last_data[user_id]['data']

        def safe_escape(text):
            if text is None:
                return ""
            return html.escape(str(text).strip())

        # Функции сортировки
        if sort_field == "price":
            def parse_price(p):
                try:
                    price_str = p.get('Цена', '0').replace('₽', '').replace(' ', '').strip()
                    return float(''.join(filter(str.isdigit, price_str)) or 0)
                except:
                    return 0.0
            products.sort(key=parse_price, reverse=(sort_order == "desc"))
        elif sort_field == "rating":
            def parse_rating(p):
                try:
                    rating_str = p.get('Рейтинг', '0').replace(',', '.').strip()
                    return float(rating_str or 0)
                except:
                    return 0.0
            products.sort(key=parse_rating, reverse=(sort_order == "desc"))
        
        # Формируем текст с результатами
        result_parts = []
        result_parts.append(f"✅ Найдено товаров: {len(products)}")
        result_parts.append(f"🔍 Поисковый запрос: {safe_escape(user_last_data[user_id].get('query', ''))}\n")
        
        for product in products:
            product_parts = []
            name = safe_escape(product.get('Название', ''))
            price = safe_escape(product.get('Цена', ''))
            rating = safe_escape(product.get('Рейтинг', '0'))
            reviews = safe_escape(product.get('Отзывы', '0'))
            link = product.get('Ссылка', '').strip()
            
            if link and not link.startswith(('http://', 'https://')):
                link = f"https://{link.lstrip('/')}"
            
            product_parts.extend([
                f"🔸 <b>{name}</b>",
                f"💲 Цена: {price}",
                f"⭐ Рейтинг: {rating}",
                f"💬 Отзывы: {reviews}"
            ])
            
            if link:
                safe_link = html.escape(link)
                product_parts.append(f"🔗 <a href=\"{safe_link}\">Ссылка на товар</a>")
            
            product_parts.append("")  # Пустая строка для разделения товаров
            result_parts.append("\n".join(product_parts))
        
        result_text = "\n".join(result_parts)
        
        # Создаем кнопки сортировки
        extra_buttons = [
            InlineKeyboardButton("По цене ⬆️", callback_data="sort_wb:price:asc"),
            InlineKeyboardButton("По цене ⬇️", callback_data="sort_wb:price:desc"),
            InlineKeyboardButton("По рейтингу ⬆️", callback_data="sort_wb:rating:asc"),
            InlineKeyboardButton("По рейтингу ⬇️", callback_data="sort_wb:rating:desc")
        ]
        
        # Сохраняем отсортированные данные
        user_last_data[user_id].update({
            'data': products,
            'text': result_text
        })
        
        # Разбиваем сообщение на части, если оно слишком длинное
        max_length = 4000
        messages = []
        current_message = ""
        
        for line in result_text.split('\n'):
            if len(current_message) + len(line) + 1 > max_length:
                messages.append(current_message)
                current_message = line + '\n'
            else:
                current_message += line + '\n'
        
        if current_message:
            messages.append(current_message)
        
        # Отправляем каждую часть отдельно
        try:
            await callback_query.message.delete()
        except:
            pass
            
        for i, msg_text in enumerate(messages):
            if i == len(messages) - 1:  # Последнее сообщение с кнопками
                await send_response(callback_query.message.chat.id, msg_text, add_download_buttons=True, action="parse_wb", extra_buttons=extra_buttons)
            else:
                await bot.send_message(callback_query.message.chat.id, msg_text, parse_mode='HTML')
        
        await callback_query.answer("Сортировка выполнена.")
    except Exception as e:
        logger.error(f"Ошибка сортировки: {e}")
        await callback_query.answer("Ошибка при сортировке.")

def format_hh_vacancy(vacancy):
    """Форматирует данные вакансии в безопасный HTML-текст"""
    try:
        def sanitize_text(text, is_url=False):
            if text is None:
                return ""
            text = str(text).strip()
            # URLs should only be escaped for HTML attributes
            if is_url:
                return text.replace('"', '%22').replace("'", '%27').replace('<', '%3C').replace('>', '%3E')
            # Regular text should be HTML escaped
            return html.escape(text)

        # Безопасное получение значений с HTML-экранированием
        def safe_get(key, default=''):
            value = vacancy.get(key)
            if value is None:
                return default
            return sanitize_text(value)

        # Формирование зарплаты
        salary_from = vacancy.get('зарплата_от')
        salary_to = vacancy.get('зарплата_до')
        currency = safe_get('валюта')
        
        salary_text = "не указана"
        if salary_from or salary_to:
            parts = []
            if salary_from:
                parts.append(f"от {salary_from}")
            if salary_to:
                parts.append(f"до {salary_to}")
            if parts:
                salary_text = f"{' '.join(parts)} {currency}"

        # Формирование URL
        url = vacancy.get('url', '')
        if isinstance(url, str):
            url = url.strip()
        else:
            url = ''
            
        url_text = ''
        if url:
            safe_url = sanitize_text(url, is_url=True)
            url_text = f'🔗 <a href="{safe_url}">Ссылка на вакансию</a>'
        else:
            url_text = "🔗 Ссылка отсутствует"

        # Получаем и очищаем все значения заранее
        name = safe_get('название', 'Название не указано')
        employer = safe_get('работодатель', 'Не указан')
        experience = safe_get('опыт', 'Не указан')
        schedule = safe_get('график_работы', 'Не указан')
        pub_date = safe_get('дата публикации', 'Не указана')

        # Собираем части сообщения с правильным экранированием
        message_parts = [
            f"🔸 <b>{name}</b>",
            f"🏢 Работодатель: {employer}",
            f"💰 Зарплата: {sanitize_text(salary_text)}",
            f"📅 Опыт: {experience}",
            f"🕓 График: {schedule}",
            url_text,
            f"📅 Дата публикации: {pub_date}",
            ""  # Пустая строка для разделения вакансий
        ]

        # Собираем все части с переносами строк
        return "\n".join(filter(None, message_parts))

    except Exception as e:
        logger.error(f"Ошибка форматирования вакансии: {e}")
        return "⚠️ Ошибка отображения вакансии\n"

@dp.message_handler(lambda message: message.from_user.id in user_states and user_states[message.from_user.id].get("action") == "parse_hh", content_types=['text'])
async def handle_hh_request(message: types.Message):
    user_id = message.from_user.id
    allowed, err_msg = db_helper.check_request_limit(user_id)
    if not allowed:
        await message.reply(err_msg)
        return

    args = message.text.strip().split(";")
    if len(args) < 2:
        await message.reply("Используйте формат: <поиск>;<город>")
        return

    query, city = args[0].strip(), args[1].strip()

    try:
        # Отправляем начальное сообщение о загрузке
        loading_message = await message.reply(
            "🔄 Начинаю поиск вакансий на HH.ru...\n"
            "⬜️⬜️⬜️⬜️⬜️ 0%"
        )

        # Создаем парсер и получаем ID города
        parser = HHParserApp()
        
        # Обновляем статус: Поиск города
        await loading_message.edit_text(
            "🔄 Определение региона поиска...\n"
            "🟩⬜️⬜️⬜️⬜️ 20%"
        )
        
        area_id = parser.get_area_id(city)
        if not area_id:
            await loading_message.edit_text("❌ Город не найден!")
            return

        # Проверяем подписку пользователя для определения лимита
        user = db_helper.get_user(user_id)
        subscription = user[0] if user else 0
        max_results = 20 if subscription == 1 else 5

        # Обновляем статус: Поиск вакансий
        await loading_message.edit_text(
            "🔄 Поиск вакансий...\n"
            "🟩🟩⬜️⬜️⬜️ 40%"
        )

        # Модифицируем метод run_parser, чтобы он принимал max_results
        parser.max_results = max_results
        vacancies = await asyncio.get_event_loop().run_in_executor(
            None, parser.run_parser, query, area_id, None, None, None
        )

        if not vacancies:
            await loading_message.edit_text("❌ Вакансий по вашему запросу не найдено")
            return

        # Обновляем статус: Обработка результатов
        await loading_message.edit_text(
            "🔄 Обработка найденных вакансий...\n"
            "🟩🟩🟩⬜️⬜️ 60%"
        )

        # Обновляем статус: Форматирование
        await loading_message.edit_text(
            "🔄 Форматирование результатов...\n"
            "🟩🟩🟩🟩⬜️ 80%"
        )

        # Формируем заголовок
        header = (
            f"✅ Найдено вакансий: {len(vacancies)}\n"
            f"🔍 Поисковый запрос: {html.escape(query)}\n"
            f"🏙 Город: {html.escape(city)}\n\n"
        )

        # Разбиваем вакансии на страницы
        VACANCIES_PER_PAGE = 3
        pages = []
        current_page = []
        current_length = len(header)

        for vacancy in vacancies:
            vacancy_text = format_hh_vacancy(vacancy)
            # Добавляем вакансию на текущую страницу
            if len(current_page) < VACANCIES_PER_PAGE:
                current_page.append(vacancy_text)
            # Если страница заполнена, сохраняем её и начинаем новую
            if len(current_page) == VACANCIES_PER_PAGE:
                pages.append(header + "\n".join(current_page))
                current_page = []

        # Добавляем оставшиеся вакансии на последнюю страницу
        if current_page:
            pages.append(header + "\n".join(current_page))

        # Если нет страниц (на случай ошибки), создаем одну пустую
        if not pages:
            pages = [header]

        # Генерируем уникальный ID для этой сессии пагинации
        global pagination_counter
        session_id = f"p_{pagination_counter}"
        pagination_counter += 1

        # Создаем клавиатуру с кнопками навигации
        keyboard = build_pagination_keyboard(session_id, 0, len(pages))

        # Сохраняем данные пагинации
        pagination_sessions[session_id] = {
            'pages': pages,
            'current_page': 0,
            'chat_id': message.chat.id,
            'message_id': None,  # Будет установлено после отправки сообщения
            'extra_buttons': [
                InlineKeyboardButton("📥 Скачать результаты:", callback_data="ignore"),
                InlineKeyboardButton("TXT", callback_data=f"download:parse_hh:txt"),
                InlineKeyboardButton("CSV", callback_data=f"download:parse_hh:csv"),
                InlineKeyboardButton("XLSX", callback_data=f"download:parse_hh:xlsx")
            ]
        }

        # Удаляем сообщение о загрузке
        await loading_message.delete()

        # Отправляем первую страницу
        msg = await bot.send_message(
            message.chat.id,
            pages[0],
            parse_mode='HTML',
            reply_markup=keyboard
        )

        # Сохраняем ID сообщения
        pagination_sessions[session_id]['message_id'] = msg.message_id

        # Сохраняем данные для возможности выгрузки
        user_last_data[user_id] = {
            'action': 'parse_hh',
            'data': vacancies,
            'text': "\n\n".join(pages),
            'hh_params': {
                'query': query,
                'city': city
            }
        }

        db_helper.add_history(user_id, 'parse_hh', query)

    except Exception as e:
        logger.error(f"Ошибка: {e}")
        if 'loading_message' in locals():
            await loading_message.edit_text("⚠️ Произошла ошибка при выполнении парсинга HH.ru")
        else:
            await message.reply("⚠️ Произошла ошибка при выполнении парсинга HH.ru")

@dp.message_handler(lambda message: message.from_user.id in user_states and user_states[message.from_user.id].get("action") == "parser_yt", content_types=['text'])
async def handle_youtube_request(message: types.Message):
    user_id = message.from_user.id
    allowed, err_msg = db_helper.check_request_limit(user_id)
    if not allowed:
        await message.reply(err_msg)
        return
    
    query = message.text.strip()
    if not query:
        await message.reply("Запрос не может быть пустым!")
        return

    try:
        # Отправляем начальное сообщение о загрузке
        loading_message = await message.reply(
            "🔄 Начинаю поиск видео на YouTube...\n"
            "⬜️⬜️⬜️⬜️⬜️ 0%"
        )

        # Удаляем сообщение пользователя для чистоты чата
        try:
            await message.delete()
        except:
            pass

        # Обновляем статус: Подключение к API
        await loading_message.edit_text(
            "🔄 Подключение к YouTube API...\n"
            "🟩⬜️⬜️⬜️⬜️ 20%"
        )

        # Проверяем подписку пользователя
        user = db_helper.get_user(user_id)
        subscription = user[0] if user else 0
        # Если есть подписка - 20 видео, если нет - 5
        max_results = 20 if subscription == 1 else 5
        
        # Обновляем статус: Поиск видео
        await loading_message.edit_text(
            "🔄 Поиск видео...\n"
            "🟩🟩🟩⬜️⬜️ 60%"
        )
        
        videos = YouTubeParser.search_videos(query, max_results=max_results)
        if not videos:
            await loading_message.edit_text("❌ Видео по вашему запросу не найдены")
            return

        # Обновляем статус: Форматирование
        await loading_message.edit_text(
            "🔄 Форматирование результатов...\n"
            "🟩🟩🟩🟩⬜️ 80%"
        )

        result_text = f"✅ Найдено видео по запросу: {html.escape(query)}\n\n"
        for idx, video in enumerate(videos, start=1):
            result_text += f"{idx}. <a href='{html.escape(video.watch_url)}'>{html.escape(video.title)}</a>\n"

        user_last_data[user_id] = {'action': 'parser_yt', 'data': videos, 'text': result_text}
        db_helper.add_history(user_id, 'parser_yt', query)

        # Создаем кнопки для скачивания
        extra_buttons = [
            InlineKeyboardButton("📥 Скачать результаты:", callback_data="ignore"),
            InlineKeyboardButton("TXT", callback_data="download:parser_yt:txt"),
            InlineKeyboardButton("CSV", callback_data="download:parser_yt:csv"),
            InlineKeyboardButton("XLSX", callback_data="download:parser_yt:xlsx")
        ]

        # Отправляем результаты с кнопками
        await loading_message.edit_text(
            result_text,
            parse_mode='HTML',
            reply_markup=build_pagination_keyboard("dummy", 0, 1, extra_buttons),
            disable_web_page_preview=True
        )

    except Exception as e:
        logger.error(f"Ошибка при поиске видео: {e}")
        if 'loading_message' in locals():
            await loading_message.edit_text("⚠️ Произошла ошибка при выполнении поиска видео")
        else:
            await message.reply("⚠️ Произошла ошибка при выполнении поиска видео")

@dp.message_handler(lambda message: message.from_user.id in user_states and user_states[message.from_user.id].get("action") == "weather", content_types=['text'])
async def handle_weather_request(message: types.Message):
    user_id = message.from_user.id
    allowed, err_msg = db_helper.check_request_limit(user_id)
    if not allowed:
        await message.reply(err_msg)
        return

    city = message.text.strip()
    if not city:
        await message.reply("Пожалуйста, укажите город!")
        return

    try:
        # Отправляем начальное сообщение о загрузке
        loading_message = await message.reply(
            "🔄 Получаю информацию о погоде...\n"
            "⬜️⬜️⬜️⬜️⬜️ 0%"
        )

        # Удаляем сообщение пользователя для чистоты чата
        try:
            await message.delete()
        except:
            pass

        # Обновляем статус: Поиск города
        await loading_message.edit_text(
            "🔄 Поиск города...\n"
            "🟩🟩⬜️⬜️⬜️ 40%"
        )

        weather_info = WeatherParser.get_weather(city)
        if not weather_info:
            await loading_message.edit_text("❌ Не удалось получить данные о погоде")
            return

        # Обновляем статус: Форматирование
        await loading_message.edit_text(
            "🔄 Форматирование результатов...\n"
            "🟩🟩🟩🟩⬜️ 80%"
        )

        result_text = (
            f"🌍 Погода в городе {html.escape(city)}:\n\n"
            f"🌡 Температура: {weather_info['temp_C']}°C\n"
            f"💨 Ветер: {weather_info['wind_speed']} км/ч\n"
            f"💧 Влажность: {weather_info['humidity']}%\n"
            f"☁️ Облачность: {weather_info['cloudcover']}%\n"
        )

        user_last_data[message.from_user.id] = {'action': 'weather', 'data': weather_info, 'text': result_text}
        db_helper.add_history(user_id, 'weather', city)

        # Создаем кнопки для скачивания
        extra_buttons = [
            InlineKeyboardButton("📥 Скачать результаты:", callback_data="ignore"),
            InlineKeyboardButton("TXT", callback_data="download:weather:txt"),
            InlineKeyboardButton("CSV", callback_data="download:weather:csv"),
            InlineKeyboardButton("XLSX", callback_data="download:weather:xlsx")
        ]

        # Отправляем результаты с кнопками
        await loading_message.edit_text(
            result_text,
            parse_mode='HTML',
            reply_markup=build_pagination_keyboard("dummy", 0, 1, extra_buttons)
        )

    except Exception as e:
        logger.error(f"Ошибка при получении погоды: {e}")
        if 'loading_message' in locals():
            await loading_message.edit_text("⚠️ Произошла ошибка при получении данных о погоде")
        else:
            await message.reply("⚠️ Произошла ошибка при получении данных о погоде")

@dp.message_handler(lambda message: message.from_user.id in user_states and user_states[message.from_user.id].get("action") == "parse_wb", content_types=['text'])
async def handle_wb_request(message: types.Message):
    user_id = message.from_user.id
    allowed, err_msg = db_helper.check_request_limit(user_id)
    if not allowed:
        await message.reply(err_msg)
        return

    query = message.text.strip()
    if not query:
        await message.reply("Запрос не может быть пустым!")
        return

    try:
        # Отправляем начальное сообщение о загрузке
        loading_message = await message.reply(
            "🔄 Начинаю поиск товаров на Wildberries...\n"
            "⬜️⬜️⬜️⬜️⬜️ 0%"
        )

        # Удаляем сообщение пользователя для чистоты чата
        try:
            await message.delete()
        except:
            pass

        # Создаем и запускаем парсер
        parser = ParserWB()
        
        # Обновляем статус: Инициализация браузера
        await loading_message.edit_text(
            "🔄 Инициализация браузера...\n"
            "🟩⬜️⬜️⬜️⬜️ 20%"
        )
        
        # Запускаем парсинг
        products = await asyncio.get_event_loop().run_in_executor(None, parser.parse, query)

        if not products:
            await loading_message.edit_text("❌ Товары по вашему запросу не найдены")
            return

        # Обновляем статус: Обработка результатов
        await loading_message.edit_text(
            "🔄 Обработка найденных товаров...\n"
            "🟩🟩🟩⬜️⬜️ 60%"
        )

        # Проверяем подписку пользователя
        user = db_helper.get_user(user_id)
        subscription = user[0] if user else 0
        max_results = 20 if subscription == 1 else 5
        products = products[:max_results]

        def safe_escape(text):
            if text is None:
                return ""
            return html.escape(str(text).strip())

        # Обновляем статус: Форматирование данных
        await loading_message.edit_text(
            "🔄 Форматирование результатов...\n"
            "🟩🟩🟩🟩⬜️ 80%"
        )

        # Формируем текст с результатами
        result_parts = []
        result_parts.append(f"✅ Найдено товаров: {len(products)}")
        result_parts.append(f"🔍 Поисковый запрос: {safe_escape(query)}\n")

        for product in products:
            product_parts = []
            name = safe_escape(product.get('Название', ''))
            price = safe_escape(product.get('Цена', ''))
            rating = safe_escape(product.get('Рейтинг', '0'))
            reviews = safe_escape(product.get('Отзывы', '0'))
            link = product.get('Ссылка', '').strip()
            
            if link and not link.startswith(('http://', 'https://')):
                link = f"https://{link.lstrip('/')}"
            
            product_parts.extend([
                f"🔸 <b>{name}</b>",
                f"💲 Цена: {price}",
                f"⭐ Рейтинг: {rating}",
                f"💬 Отзывы: {reviews}"
            ])
            
            if link:
                safe_link = html.escape(link)
                product_parts.append(f"🔗 <a href=\"{safe_link}\">Ссылка на товар</a>")
            
            product_parts.append("")  # Пустая строка для разделения товаров
            result_parts.append("\n".join(product_parts))

        result_text = "\n".join(result_parts)

        # Создаем кнопки сортировки и скачивания
        extra_buttons = [
            InlineKeyboardButton("По цене ⬆️", callback_data="sort_wb:price:asc"),
            InlineKeyboardButton("По цене ⬇️", callback_data="sort_wb:price:desc"),
            InlineKeyboardButton("По рейтингу ⬆️", callback_data="sort_wb:rating:asc"),
            InlineKeyboardButton("По рейтингу ⬇️", callback_data="sort_wb:rating:desc"),
            InlineKeyboardButton("📥 Скачать результаты:", callback_data="ignore"),
            InlineKeyboardButton("TXT", callback_data="download:parse_wb:txt"),
            InlineKeyboardButton("CSV", callback_data="download:parse_wb:csv"),
            InlineKeyboardButton("XLSX", callback_data="download:parse_wb:xlsx")
        ]

        # Сохраняем данные
        user_last_data[user_id] = {
            'action': 'parse_wb',
            'data': products,
            'text': result_text,
            'query': query
        }

        db_helper.add_history(user_id, 'parse_wb', query)

        # Отправляем результаты с кнопками
        await loading_message.edit_text(
            result_text,
            parse_mode='HTML',
            reply_markup=build_pagination_keyboard("dummy", 0, 1, extra_buttons),
            disable_web_page_preview=True
        )

    except Exception as e:
        logger.error(f"Ошибка при парсинге Wildberries: {e}")
        if 'loading_message' in locals():
            await loading_message.edit_text("⚠️ Произошла ошибка при поиске товаров на Wildberries")
        else:
            await message.reply("⚠️ Произошла ошибка при поиске товаров на Wildberries")

@dp.callback_query_handler(lambda c: c.data.startswith("download:"))
async def process_download_callback(callback_query: types.CallbackQuery):
    try:
        data = callback_query.data.split(":")  # формат: download:<action>:<format>
        if len(data) != 3:
            await callback_query.answer("Некорректные данные.")
            return
        
        action, file_format = data[1], data[2]
        user_id = callback_query.from_user.id
        
        if user_id not in user_last_data:
            await callback_query.message.reply("❌ Данные устарели. Пожалуйста, выполните поиск заново.")
            return
            
        result = user_last_data[user_id]
        buf = BytesIO()
        filename = f"result_{action}.{file_format}"

        if file_format == "xlsx":
            wb = Workbook()
            ws = wb.active
            header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            header_font = Font(bold=True)

            if action == "parser_yt":
                headers = ["№", "Название", "Ссылка"]
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col, value=header)
                    cell.fill = header_fill
                    cell.font = header_font
                for idx, video in enumerate(result['data'], 2):
                    ws.cell(row=idx, column=1, value=idx-1)
                    ws.cell(row=idx, column=2, value=video.title)
                    ws.cell(row=idx, column=3, value=video.watch_url)

            elif action == "parse_hh":
                headers = ["№", "Название", "Работодатель", "Зарплата от", "Зарплата до", "Валюта", "Опыт", "График работы", "Ссылка", "Дата публикации"]
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col, value=header)
                    cell.fill = header_fill
                    cell.font = header_font
                for idx, vac in enumerate(result['data'], 2):
                    ws.cell(row=idx, column=1, value=idx-1)
                    ws.cell(row=idx, column=2, value=vac.get('название'))
                    ws.cell(row=idx, column=3, value=vac.get('работодатель'))
                    ws.cell(row=idx, column=4, value=vac.get('зарплата_от'))
                    ws.cell(row=idx, column=5, value=vac.get('зарплата_до'))
                    ws.cell(row=idx, column=6, value=vac.get('валюта'))
                    ws.cell(row=idx, column=7, value=vac.get('опыт'))
                    ws.cell(row=idx, column=8, value=vac.get('график_работы'))
                    ws.cell(row=idx, column=9, value=vac.get('url'))
                    ws.cell(row=idx, column=10, value=vac.get('дата публикации'))

            elif action == "parse_wb":
                headers = ["№", "Название", "Цена", "Рейтинг", "Отзывы", "Ссылка"]
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col, value=header)
                    cell.fill = header_fill
                    cell.font = header_font
                for idx, prod in enumerate(result['data'], 2):
                    ws.cell(row=idx, column=1, value=idx-1)
                    ws.cell(row=idx, column=2, value=prod.get('Название'))
                    ws.cell(row=idx, column=3, value=prod.get('Цена'))
                    ws.cell(row=idx, column=4, value=prod.get('Рейтинг'))
                    ws.cell(row=idx, column=5, value=prod.get('Отзывы'))
                    ws.cell(row=idx, column=8, value=prod.get('Ссылка'))

            elif action == "weather":
                headers = ["Параметр", "Значение"]
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col, value=header)
                    cell.fill = header_fill
                    cell.font = header_font
                weather_data = result['data']
                row = 2
                for key, value in weather_data.items():
                    ws.cell(row=row, column=1, value=key)
                    ws.cell(row=row, column=2, value=value)
                    row += 1

            # Автоматическая настройка ширины столбцов
            for column in ws.columns:
                max_length = 0
                column = list(column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column[0].column_letter].width = adjusted_width

            wb.save(buf)
            filename = f"result_{action}.xlsx"

        elif file_format == "csv":
            output = StringIO()
            csv_writer = csv.writer(output)
            if action == "parser_yt":
                csv_writer.writerow(["№", "Название", "Ссылка"])
                for idx, video in enumerate(result['data'], start=1):
                    csv_writer.writerow([idx, video.title, video.watch_url])
            elif action == "parse_hh":
                csv_writer.writerow(["№", "Название", "Работодатель", "Зарплата от", "Зарплата до", "Валюта", "Опыт", "График работы", "Ссылка", "Дата публикации"])
                for idx, vac in enumerate(result['data'], start=1):
                    csv_writer.writerow([idx, vac.get('название'), vac.get('работодатель'), vac.get('зарплата_от'), vac.get('зарплата_до'),
                                         vac.get('валюта'), vac.get('опыт'), vac.get('график_работы'), vac.get('url'), vac.get('дата публикации')])
            elif action == "parse_wb":
                csv_writer.writerow(["№", "Название", "Цена", "Рейтинг", "Отзывы", "Ссылка"])
                for idx, prod in enumerate(result['data'], start=1):
                    csv_writer.writerow([idx, prod.get('Название'), prod.get('Цена'), prod.get('Рейтинг'), prod.get('Отзывы'),
                                         prod.get('Количество продаж'), prod.get('Скидка'), prod.get('Ссылка')])
            elif action == "weather":
                weather_data = result['data']
                for key, value in weather_data.items():
                    csv_writer.writerow([key, value])
            else:
                csv_writer.writerow([result['text']])
            buf.write(output.getvalue().encode('utf-8'))

        elif file_format == "txt":
            buf.write(result['text'].encode('utf-8'))
        else:
            await callback_query.answer("Неверный формат.")
            return

        buf.seek(0)
        await bot.send_document(chat_id=callback_query.message.chat.id, document=types.InputFile(buf, filename=filename))
        await callback_query.answer("Файл отправлен.")
    except Exception as e:
        logger.error(f"Ошибка при скачивании файла: {e}")
        try:
            await callback_query.message.reply(
                "❌ Произошла ошибка при подготовке файла. Пожалуйста, попробуйте еще раз."
            )
        except:
            pass



@dp.message_handler(commands=['subpass'])
async def handle_subpass(message: types.Message):
    # Очищаем состояние пользователя
    if message.from_user.id in user_states:
        del user_states[message.from_user.id]
    
    args = message.get_args().strip()
    if args == "Kakashki":
        db_helper.update_subscription(message.from_user.id, 1)
        await message.reply("✅ Подписка успешно выдана! Теперь у вас неограниченное количество запросов.")
    else:
        await message.reply("❌ Неверный пароль!")

async def create_invoice(amount: float, currency: str, description: str) -> dict:
    url = 'https://pay.crypt.bot/api/createInvoice'
    headers = {'Crypto-Pay-API-Token': CRYPTO_PAY_API_TOKEN}
    data = {'asset': currency, 'amount': str(amount), 'description': description}
    try:
        async with aiohttp.ClientSession() as session:
            async with session.post(url, headers=headers, json=data) as response:
                response.raise_for_status()
                result = await response.json()
                logger.info(f"Ответ API: status_code={response.status}, result={result}")
                return result
    except aiohttp.ClientError as e:
        logger.error(f"Ошибка при создании счёта: {e}")
        return None

async def check_invoice_status(invoice_id: str) -> str:
    url = f'https://pay.crypt.bot/api/getInvoices?invoice_ids={invoice_id}'
    headers = {'Crypto-Pay-API-Token': CRYPTO_PAY_API_TOKEN}
    try:
        async with aiohttp.ClientSession() as session:
            async with session.get(url, headers=headers) as response:
                response.raise_for_status()
                data = await response.json()
                if data.get('result') and len(data['result']) > 0:
                    return data['result'][0].get('status', 'unknown')
                return "not_found"
    except aiohttp.ClientError as e:
        logger.error(f"Ошибка при проверке статуса счёта: {e}")
        return "error"

@dp.message_handler(commands=['subscribe'])
async def handle_subscribe(message: types.Message):
    amount = 1  # стоимость подписки в USD
    currency = 'USDT'
    description = "Подписка на бота: неограниченные запросы"
    invoice_data = await create_invoice(amount, currency, description)
    if invoice_data and invoice_data.get('result'):
        payment_url = invoice_data['result'].get('pay_url')
        invoice_id = invoice_data['result'].get('invoice_id')
        keyboard = InlineKeyboardMarkup(row_width=2)
        keyboard.add(
            InlineKeyboardButton("💳 Оплатить подписку", url=payment_url),
            InlineKeyboardButton("✅ Проверить оплату", callback_data=f"check_payment:{invoice_id}")
        )
        await message.reply(
            "Для оплаты подписки нажмите кнопку ниже:\n"
            "После успешной оплаты нажмите 'Проверить оплату'",
            reply_markup=keyboard
        )
    else:
        await message.reply("Ошибка при создании счёта для оплаты. Попробуйте позже.")

@dp.callback_query_handler(lambda c: c.data.startswith('check_payment:'))
async def process_callback(callback_query: types.CallbackQuery):
    invoice_id = callback_query.data.split(':')[1]
    status = await check_invoice_status(invoice_id)
    if status == "paid":
        db_helper.update_subscription(callback_query.from_user.id, 1)
        await bot.answer_callback_query(
            callback_query.id,
            text="Оплата подтверждена! Подписка активирована.",
            show_alert=True
        )
        await bot.edit_message_reply_markup(
            chat_id=callback_query.message.chat.id,
            message_id=callback_query.message.message_id,
            reply_markup=None
        )
    elif status == "not_found":
        await bot.answer_callback_query(
            callback_query.id,
            text="Счет не найден. Попробуйте создать новый.",
            show_alert=True
        )
    else:
        await bot.answer_callback_query(
            callback_query.id,
            text=f"Статус оплаты: {status}. Попробуйте позже.",
            show_alert=True
        )

@dp.message_handler(commands=['confirm'])
async def handle_confirm(message: types.Message):
    args = message.get_args().strip()
    if not args:
        await message.reply("Укажите invoice_id, полученный после оплаты, пример: /confirm 12345")
        return
    invoice_id = args
    status = await check_invoice_status(invoice_id)
    if status == "paid":
        db_helper.update_subscription(message.from_user.id, 1)
        await message.reply("Оплата подтверждена! Подписка оформлена. Ограничения сняты.")
    else:
        await message.reply(f"Статус счёта: {status}. Оплата не завершена или еще не подтверждена.")

@dp.callback_query_handler(lambda c: c.data.startswith("download_video:"))
async def download_video_callback(callback_query: types.CallbackQuery):
    index_str = callback_query.data.split(":")[1]
    try:
        index = int(index_str)
    except ValueError:
        await callback_query.answer("Некорректный номер видео.")
        return

    user_id = callback_query.from_user.id
    if user_id not in user_last_data or user_last_data[user_id].get('action') != 'parser_yt':
        await callback_query.answer("Нет данных для скачивания видео.")
        return

    videos = user_last_data[user_id]['data']
    if index < 0 or index >= len(videos):
        await callback_query.answer("Видео с таким номером не найдено.")
        return

    video = videos[index]
    try:
        file_path = YouTubeParser.download_video(video.watch_url)
        if file_path:
            await bot.send_video(callback_query.message.chat.id, video=open(file_path, 'rb'))
            YouTubeParser.cleanup_file(file_path)
            await callback_query.answer("Видео отправлено.")
        else:
            await callback_query.answer("Не удалось скачать видео.")
    except Exception as e:
        logger.error(f"Ошибка скачивания видео: {e}")
        await callback_query.answer("Ошибка при скачивании видео.")

def get_download_keyboard(action):
    keyboard = InlineKeyboardMarkup()
    keyboard.row(
        InlineKeyboardButton("CSV", callback_data=f"download:{action}:csv"),
        InlineKeyboardButton("TXT", callback_data=f"download:{action}:txt"),
        InlineKeyboardButton("Excel", callback_data=f"download:{action}:xlsx")
    )
    return keyboard

# Состояния диалога для HH.ru
HH_STATES = {
    'waiting_job': 'Ожидание ввода профессии',
    'waiting_city': 'Ожидание ввода города',
    'waiting_salary': 'Ожидание ввода зарплаты',
    'waiting_work_type': 'Ожидание выбора формата работы'
}

@dp.callback_query_handler(lambda c: c.data == "parse_hh")
async def start_hh_dialog(callback_query: types.CallbackQuery):
    user_id = callback_query.from_user.id
    # Отправляем первое сообщение и сохраняем его ID
    msg = await bot.send_message(
        user_id,
        "💼 Давайте найдем подходящую работу!\n\n"
        "Какую работу ищете? Введите название профессии или должности:"
    )
    user_states[user_id] = {
        "action": "parse_hh",
        "state": "waiting_job",
        "data": {},
        "dialog_message_id": msg.message_id  # Сохраняем ID сообщения
    }
    await callback_query.answer()

@dp.message_handler(lambda message: message.from_user.id in user_states and 
                   user_states[message.from_user.id].get("action") == "parse_hh" and 
                   "state" not in user_states[message.from_user.id])
async def handle_old_hh_format(message: types.Message):
    # Перенаправляем на новый формат диалога
    user_id = message.from_user.id
    msg = await bot.send_message(
        user_id,
        "💼 Теперь поиск работы стал удобнее! Давайте начнем сначала.\n\n"
        "Какую работу ищете? Введите название профессии или должности:"
    )
    user_states[user_id] = {
        "action": "parse_hh",
        "state": "waiting_job",
        "data": {},
        "dialog_message_id": msg.message_id
    }

@dp.message_handler(lambda message: message.from_user.id in user_states and 
                   user_states[message.from_user.id].get("action") == "parse_hh" and 
                   user_states[message.from_user.id].get("state") in HH_STATES)
async def handle_hh_dialog(message: types.Message):
    user_id = message.from_user.id
    current_state = user_states[user_id]["state"]
    user_data = user_states[user_id].get("data", {})
    dialog_message_id = user_states[user_id].get("dialog_message_id")

    # Удаляем сообщение пользователя для чистоты чата
    try:
        await message.delete()
    except:
        pass

    if current_state == "waiting_job":
        user_data["job"] = message.text.strip()
        user_states[user_id]["state"] = "waiting_city"
        await bot.edit_message_text(
            "✅ Профессия: " + user_data["job"] + "\n\n"
            "В каком городе ищете работу?",
            chat_id=user_id,
            message_id=dialog_message_id
        )

    elif current_state == "waiting_city":
        user_data["city"] = message.text.strip()
        user_states[user_id]["state"] = "waiting_salary"
        await bot.edit_message_text(
            "✅ Профессия: " + user_data["job"] + "\n"
            "✅ Город: " + user_data["city"] + "\n\n"
            "Укажите желаемую зарплату (или напишите 'любая'):",
            chat_id=user_id,
            message_id=dialog_message_id
        )

    elif current_state == "waiting_salary":
        salary_text = message.text.strip().lower()
        if salary_text != "любая":
            try:
                salary = int(salary_text)
                user_data["salary"] = salary
            except ValueError:
                await bot.edit_message_text(
                    "✅ Профессия: " + user_data["job"] + "\n"
                    "✅ Город: " + user_data["city"] + "\n"
                    "❌ Ошибка: введите число или слово 'любая'\n\n"
                    "Укажите желаемую зарплату:",
                    chat_id=user_id,
                    message_id=dialog_message_id
                )
                return
        else:
            user_data["salary"] = None

        # Создаем клавиатуру для выбора формата работы
        keyboard = InlineKeyboardMarkup(row_width=1)
        keyboard.add(
            InlineKeyboardButton("🏢 Офис", callback_data="work_type:office"),
            InlineKeyboardButton("🏠 Удаленная работа", callback_data="work_type:remote"),
            InlineKeyboardButton("🔄 Гибридный формат", callback_data="work_type:hybrid"),
            InlineKeyboardButton("📋 Любой формат", callback_data="work_type:any")
        )
        user_states[user_id]["state"] = "waiting_work_type"
        
        salary_display = "любая" if user_data["salary"] is None else str(user_data["salary"])
        await bot.edit_message_text(
            "✅ Профессия: " + user_data["job"] + "\n"
            "✅ Город: " + user_data["city"] + "\n"
            "✅ Зарплата: " + salary_display + "\n\n"
            "Выберите предпочтительный формат работы:",
            chat_id=user_id,
            message_id=dialog_message_id,
            reply_markup=keyboard
        )

    user_states[user_id]["data"] = user_data

@dp.callback_query_handler(lambda c: c.data.startswith("work_type:"))
async def process_work_type(callback_query: types.CallbackQuery):
    try:
        user_id = callback_query.from_user.id
        if user_id not in user_states or user_states[user_id].get("action") != "parse_hh":
            await callback_query.message.edit_text(
                "❌ Сессия поиска устарела. Пожалуйста, начните поиск заново.",
                reply_markup=None
            )
            return

        work_type = callback_query.data.split(":")[1]
        user_data = user_states[user_id].get("data", {})
        user_data["work_type"] = work_type

        # Обновляем сообщение, показывая все выбранные параметры
        work_type_text = {
            "office": "Офис",
            "remote": "Удаленная работа",
            "hybrid": "Гибридный формат",
            "any": "Любой формат"
        }.get(work_type, "Не указан")

        salary_display = "любая" if user_data.get("salary") is None else str(user_data["salary"])
        await callback_query.message.edit_text(
            "✅ Профессия: " + user_data["job"] + "\n"
            "✅ Город: " + user_data["city"] + "\n"
            "✅ Зарплата: " + salary_display + "\n"
            "✅ Формат работы: " + work_type_text + "\n\n"
            "🔄 Начинаю поиск вакансий..."
        )

        # Начинаем поиск вакансий
        await start_hh_search(callback_query.message, user_id, user_data)
        await callback_query.answer()
    except Exception as e:
        logger.error(f"Ошибка в обработчике выбора типа работы: {e}")
        try:
            await callback_query.message.edit_text(
                "❌ Произошла ошибка. Пожалуйста, начните поиск заново.",
                reply_markup=None
            )
        except:
            pass

async def start_hh_search(message, user_id, user_data):
    allowed, err_msg = db_helper.check_request_limit(user_id)
    if not allowed:
        await message.reply(err_msg)
        return

    try:
        # Отправляем начальное сообщение о загрузке
        loading_message = await message.reply(
            "🔄 Начинаю поиск вакансий на HH.ru...\n"
            "⬜️⬜️⬜️⬜️⬜️ 0%"
        )

        # Создаем парсер и получаем ID города
        parser = HHParserApp()
        
        # Обновляем статус: Поиск города
        await loading_message.edit_text(
            "🔄 Определение региона поиска...\n"
            "🟩⬜️⬜️⬜️⬜️ 20%"
        )
        
        area_id = parser.get_area_id(user_data["city"])
        if not area_id:
            await loading_message.edit_text("❌ Город не найден!")
            return

        # Проверяем подписку пользователя для определения лимита
        user = db_helper.get_user(user_id)
        subscription = user[0] if user else 0
        max_results = 20 if subscription == 1 else 5

        # Обновляем статус: Поиск вакансий
        await loading_message.edit_text(
            "🔄 Поиск вакансий...\n"
            "🟩🟩⬜️⬜️⬜️ 40%"
        )

        # Определяем параметры поиска
        remote_filter = None
        if user_data["work_type"] == "remote":
            remote_filter = "Удаленная работа"

        # Модифицируем метод run_parser, чтобы он принимал max_results
        parser.max_results = max_results
        vacancies = await asyncio.get_event_loop().run_in_executor(
            None, parser.run_parser, 
            user_data["job"], 
            area_id,
            user_data.get("salary"),
            None,
            remote_filter
        )

        if not vacancies:
            await loading_message.edit_text("❌ Вакансий по вашему запросу не найдено")
            return

        # Обновляем статус: Форматирование
        await loading_message.edit_text(
            "🔄 Форматирование результатов...\n"
            "🟩🟩🟩🟩⬜️ 80%"
        )

        # Формируем заголовок
        work_type_text = {
            "office": "Офис",
            "remote": "Удаленная работа",
            "hybrid": "Гибридный формат",
            "any": "Любой формат"
        }.get(user_data["work_type"], "Не указан")

        header = (
            f"✅ Найдено вакансий: {len(vacancies)}\n"
            f"🔍 Поиск: {html.escape(user_data['job'])}\n"
            f"🏙 Город: {html.escape(user_data['city'])}\n"
            f"💰 Зарплата от: {user_data.get('salary', 'не указана')}\n"
            f"🏢 Формат работы: {work_type_text}\n\n"
        )

        # Разбиваем вакансии на страницы
        VACANCIES_PER_PAGE = 3
        pages = []
        current_page = []

        for vacancy in vacancies:
            vacancy_text = format_hh_vacancy(vacancy)
            if len(current_page) < VACANCIES_PER_PAGE:
                current_page.append(vacancy_text)
            if len(current_page) == VACANCIES_PER_PAGE:
                pages.append(header + "\n".join(current_page))
                current_page = []

        if current_page:
            pages.append(header + "\n".join(current_page))

        if not pages:
            pages = [header]

        # Генерируем уникальный ID для этой сессии пагинации
        global pagination_counter
        session_id = f"p_{pagination_counter}"
        pagination_counter += 1

        # Создаем клавиатуру с кнопками навигации
        keyboard = build_pagination_keyboard(session_id, 0, len(pages))

        # Сохраняем данные пагинации
        pagination_sessions[session_id] = {
            'pages': pages,
            'current_page': 0,
            'chat_id': message.chat.id,
            'message_id': None,
            'extra_buttons': [
                InlineKeyboardButton("📥 Скачать результаты:", callback_data="ignore"),
                InlineKeyboardButton("TXT", callback_data=f"download:parse_hh:txt"),
                InlineKeyboardButton("CSV", callback_data=f"download:parse_hh:csv"),
                InlineKeyboardButton("XLSX", callback_data=f"download:parse_hh:xlsx")
            ]
        }

        # Удаляем сообщение о загрузке
        await loading_message.delete()

        # Отправляем первую страницу
        msg = await bot.send_message(
            message.chat.id,
            pages[0],
            parse_mode='HTML',
            reply_markup=keyboard
        )

        # Сохраняем ID сообщения
        pagination_sessions[session_id]['message_id'] = msg.message_id

        # Сохраняем данные для возможности выгрузки
        user_last_data[user_id] = {
            'action': 'parse_hh',
            'data': vacancies,
            'text': "\n\n".join(pages),
            'hh_params': user_data
        }

        db_helper.add_history(user_id, 'parse_hh', user_data['job'])

    except Exception as e:
        logger.error(f"Ошибка: {e}")
        if 'loading_message' in locals():
            await loading_message.edit_text("⚠️ Произошла ошибка при выполнении парсинга HH.ru")
        else:
            await message.reply("⚠️ Произошла ошибка при выполнении парсинга HH.ru")

if __name__ == '__main__':
    loop = asyncio.get_event_loop()
    loop.create_task(subscription_checker())
    executor.start_polling(dp, skip_updates=True)
#━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━